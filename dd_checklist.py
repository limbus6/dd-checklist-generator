"""
Due Diligence Document Checklist Generator for M&A Transactions.

Generates a formatted Excel checklist based on transaction type, sector,
jurisdiction, and target company name. Supports PT/EN languages.

Usage:
    python dd_checklist.py
"""

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TRANSACTION_TYPES = ["Asset Deal", "Share Deal", "Merger"]
SECTORS = [
    "Healthcare",
    "Technology",
    "Industrial",
    "Real Estate",
    "Financial Services",
    "Retail",
]
JURISDICTIONS = ["Portugal", "Espanha", "Internacional"]

CATEGORIES = [
    "Legal",
    "Financial",
    "Operational",
    "Tax",
    "HR",
    "Commercial",
    "IP",
    "Compliance",
]

# ---------------------------------------------------------------------------
# Document definitions — bilingual (EN / PT)
# ---------------------------------------------------------------------------

# Each document: (category, name_en, name_pt, required, priority)

CORE_DOCUMENTS = [
    ("Legal", "Articles of Association / By-laws", "Estatutos / Pacto Social", "Yes", "High"),
    ("Legal", "Certificate of Incorporation", "Certidão Permanente", "Yes", "High"),
    ("Legal", "Board minutes (last 3 years)", "Atas de assembleia (últimos 3 anos)", "Yes", "High"),
    ("Legal", "Powers of Attorney in force", "Procurações em vigor", "Yes", "Medium"),
    ("Legal", "Pending / threatened litigation", "Litígios pendentes / ameaçados", "Yes", "High"),
    ("Legal", "Regulatory licences & permits", "Licenças e alvarás regulatórios", "Yes", "High"),
    ("Financial", "Audited Financial Statements (3 years)", "Demonstrações Financeiras auditadas (3 anos)", "Yes", "High"),
    ("Financial", "Management accounts (YTD)", "Balancetes de gestão (YTD)", "Yes", "High"),
    ("Financial", "Budget / Forecasts", "Orçamento / Projeções", "Yes", "Medium"),
    ("Financial", "Debt schedule & loan agreements", "Mapa de dívida e contratos de empréstimo", "Yes", "High"),
    ("Financial", "Bank statements (12 months)", "Extratos bancários (12 meses)", "Yes", "Medium"),
    ("Financial", "Accounts receivable & payable aging", "Aging de contas a receber e a pagar", "Yes", "Medium"),
    ("Tax", "Corporate tax returns (3 years)", "Declarações IRC (3 anos)", "Yes", "High"),
    ("Tax", "VAT returns (3 years)", "Declarações IVA (3 anos)", "Yes", "High"),
    ("Tax", "Tax assessments / disputes", "Avaliações / litígios fiscais", "Yes", "High"),
    ("Tax", "Transfer pricing documentation", "Documentação de preços de transferência", "No", "Medium"),
    ("HR", "Employee list with terms", "Lista de colaboradores com condições", "Yes", "High"),
    ("HR", "Employment contracts (key personnel)", "Contratos de trabalho (pessoal-chave)", "Yes", "High"),
    ("HR", "Collective bargaining agreements", "Convenções coletivas de trabalho", "Yes", "Medium"),
    ("HR", "Pension / benefit plans", "Planos de pensões / benefícios", "Yes", "Medium"),
    ("HR", "Organizational chart", "Organograma", "Yes", "Low"),
    ("Commercial", "Top 10 customer contracts", "Contratos dos 10 maiores clientes", "Yes", "High"),
    ("Commercial", "Top 10 supplier contracts", "Contratos dos 10 maiores fornecedores", "Yes", "High"),
    ("Commercial", "Material contracts summary", "Resumo de contratos materiais", "Yes", "High"),
    ("Compliance", "Data protection / GDPR policies", "Políticas de proteção de dados / RGPD", "Yes", "High"),
    ("Compliance", "Anti-money laundering policies", "Políticas de prevenção de branqueamento", "No", "Medium"),
    ("Compliance", "Insurance policies schedule", "Mapa de apólices de seguro", "Yes", "High"),
    ("Compliance", "Insurance claims history", "Histórico de sinistros", "No", "Medium"),
]


def _sector_documents():
    """Return sector-specific documents dict keyed by sector name."""
    return {
        "Healthcare": [
            ("Compliance", "Medical / healthcare operating licences", "Licenças de atividade médica / saúde", "Yes", "High"),
            ("Compliance", "Patient data compliance (GDPR health data)", "Conformidade dados de pacientes (RGPD dados de saúde)", "Yes", "High"),
            ("Operational", "Equipment certifications & calibration logs", "Certificações de equipamentos e registos de calibração", "Yes", "High"),
            ("Compliance", "Clinical trial authorizations", "Autorizações de ensaios clínicos", "No", "Medium"),
            ("Compliance", "Pharmacy / drug distribution licences", "Licenças de farmácia / distribuição de medicamentos", "No", "High"),
            ("HR", "Medical staff credentials & licences", "Credenciais e cédulas profissionais do pessoal médico", "Yes", "High"),
            ("Operational", "Health & safety inspection reports", "Relatórios de inspeção de saúde e segurança", "Yes", "Medium"),
            ("Compliance", "Agreements with national health service", "Acordos com o Serviço Nacional de Saúde", "No", "Medium"),
        ],
        "Technology": [
            ("IP", "IP portfolio (patents, trademarks, domains)", "Portfólio de PI (patentes, marcas, domínios)", "Yes", "High"),
            ("IP", "Software licence agreements (inbound)", "Contratos de licença de software (inbound)", "Yes", "High"),
            ("IP", "Software licence agreements (outbound / SaaS)", "Contratos de licença de software (outbound / SaaS)", "Yes", "High"),
            ("IP", "Source code escrow agreements", "Contratos de escrow de código-fonte", "No", "Medium"),
            ("IP", "Open source software audit", "Auditoria de software open source", "Yes", "High"),
            ("Commercial", "SaaS / subscription metrics (ARR, churn, LTV)", "Métricas SaaS / subscrição (ARR, churn, LTV)", "Yes", "High"),
            ("Operational", "IT infrastructure & security audit", "Auditoria de infraestrutura TI e segurança", "Yes", "High"),
            ("Compliance", "Data breach history & incident response plan", "Histórico de violações de dados e plano de resposta", "Yes", "Medium"),
            ("HR", "Key developer / tech talent retention plans", "Planos de retenção de talento tecnológico-chave", "No", "Medium"),
            ("Commercial", "Customer contracts with SLA details", "Contratos de clientes com detalhe de SLA", "Yes", "Medium"),
        ],
        "Industrial": [
            ("Compliance", "Environmental permits & impact assessments", "Licenças ambientais e avaliações de impacto", "Yes", "High"),
            ("Compliance", "Health & Safety certifications (ISO 45001)", "Certificações de Saúde e Segurança (ISO 45001)", "Yes", "High"),
            ("Operational", "Equipment maintenance logs", "Registos de manutenção de equipamentos", "Yes", "Medium"),
            ("Operational", "Production capacity reports", "Relatórios de capacidade produtiva", "Yes", "Medium"),
            ("Compliance", "Environmental remediation obligations", "Obrigações de remediação ambiental", "Yes", "High"),
            ("Operational", "Supply chain / logistics contracts", "Contratos de cadeia de abastecimento / logística", "Yes", "Medium"),
            ("Compliance", "Quality management certifications (ISO 9001)", "Certificações de gestão de qualidade (ISO 9001)", "Yes", "Medium"),
            ("Operational", "Fixed asset register with valuations", "Registo de ativos fixos com avaliações", "Yes", "High"),
        ],
        "Real Estate": [
            ("Legal", "Property title deeds / Certidões prediais", "Escrituras de propriedade / Certidões prediais", "Yes", "High"),
            ("Legal", "Land registry certificates", "Certidões do registo predial", "Yes", "High"),
            ("Commercial", "Lease agreements (tenant schedule)", "Contratos de arrendamento (mapa de inquilinos)", "Yes", "High"),
            ("Legal", "Building permits & occupancy licences", "Licenças de construção e utilização", "Yes", "High"),
            ("Financial", "Independent property valuations", "Avaliações independentes de imóveis", "Yes", "High"),
            ("Compliance", "Environmental site assessments", "Avaliações ambientais dos imóveis", "Yes", "Medium"),
            ("Operational", "Property management contracts", "Contratos de gestão de propriedades", "Yes", "Medium"),
            ("Financial", "Rental income schedule & vacancy rates", "Mapa de rendas e taxas de desocupação", "Yes", "High"),
            ("Legal", "Easements, encumbrances & restrictions", "Servidões, ónus e restrições", "Yes", "High"),
        ],
        "Financial Services": [
            ("Compliance", "Regulatory licences (Central Bank / CMVM / ASF)", "Licenças regulatórias (Banco de Portugal / CMVM / ASF)", "Yes", "High"),
            ("Compliance", "Capital adequacy / solvency reports", "Relatórios de adequação de capital / solvência", "Yes", "High"),
            ("Compliance", "AML / KYC policies & procedures", "Políticas e procedimentos AML / KYC", "Yes", "High"),
            ("Compliance", "Regulatory inspection reports", "Relatórios de inspeções regulatórias", "Yes", "High"),
            ("Financial", "Loan / credit portfolio analysis", "Análise da carteira de crédito", "Yes", "High"),
            ("Financial", "Provision / impairment schedules", "Mapas de provisões / imparidades", "Yes", "High"),
            ("Compliance", "Compliance officer reports (2 years)", "Relatórios do compliance officer (2 anos)", "Yes", "Medium"),
            ("Operational", "IT systems & cybersecurity audit", "Auditoria de sistemas TI e cibersegurança", "Yes", "High"),
            ("Compliance", "Client complaints register", "Registo de reclamações de clientes", "No", "Medium"),
        ],
        "Retail": [
            ("Commercial", "Franchise / distribution agreements", "Contratos de franquia / distribuição", "Yes", "High"),
            ("Commercial", "E-commerce platform details & metrics", "Detalhes e métricas da plataforma e-commerce", "No", "Medium"),
            ("Legal", "Store lease agreements", "Contratos de arrendamento de lojas", "Yes", "High"),
            ("IP", "Brand / trademark registrations", "Registos de marca", "Yes", "High"),
            ("Operational", "Inventory management reports", "Relatórios de gestão de inventário", "Yes", "Medium"),
            ("Commercial", "Loyalty programme details", "Detalhes do programa de fidelização", "No", "Low"),
            ("Compliance", "Consumer protection compliance", "Conformidade com proteção do consumidor", "Yes", "Medium"),
            ("Operational", "Store network profitability analysis", "Análise de rentabilidade da rede de lojas", "Yes", "High"),
        ],
    }


def _deal_documents():
    """Return deal-type-specific documents dict keyed by deal type."""
    return {
        "Asset Deal": [
            ("Legal", "Detailed asset list with descriptions", "Lista detalhada de ativos com descrições", "Yes", "High"),
            ("Legal", "Asset transfer agreements (drafts)", "Contratos de transferência de ativos (minutas)", "Yes", "High"),
            ("Legal", "Third-party consents for asset transfer", "Consentimentos de terceiros para transferência de ativos", "Yes", "High"),
            ("Tax", "Tax implications analysis of asset transfer", "Análise de implicações fiscais da transferência de ativos", "Yes", "High"),
            ("Financial", "Asset valuations / appraisals", "Avaliações de ativos", "Yes", "High"),
            ("Legal", "Assumed vs excluded liabilities schedule", "Mapa de passivos assumidos vs excluídos", "Yes", "High"),
        ],
        "Share Deal": [
            ("Legal", "Shareholder agreements", "Acordos parassociais", "Yes", "High"),
            ("Legal", "Share certificates", "Títulos de participação / certificados de ações", "Yes", "High"),
            ("Legal", "Capitalisation table (Cap table)", "Tabela de capitalização (Cap table)", "Yes", "High"),
            ("Legal", "Share transfer restrictions / pre-emption rights", "Restrições de transmissão de ações / direitos de preferência", "Yes", "High"),
            ("Legal", "Drag-along / tag-along provisions", "Cláusulas de drag-along / tag-along", "Yes", "Medium"),
            ("Legal", "Minority shareholder rights", "Direitos de acionistas minoritários", "Yes", "Medium"),
            ("Financial", "Dividend history & policy", "Histórico e política de dividendos", "Yes", "Medium"),
            ("Legal", "Stock option / warrant agreements", "Contratos de stock options / warrants", "No", "Medium"),
        ],
        "Merger": [
            ("Legal", "Merger plan / projeto de fusão", "Projeto de fusão", "Yes", "High"),
            ("Financial", "Fairness opinion", "Fairness opinion", "Yes", "High"),
            ("Legal", "Exchange ratio justification", "Fundamentação da relação de troca", "Yes", "High"),
            ("Legal", "Merger filing / regulatory notifications", "Notificações regulatórias da fusão", "Yes", "High"),
            ("Compliance", "Competition / antitrust analysis", "Análise concorrencial / antitrust", "Yes", "High"),
            ("Legal", "Creditor notification process documentation", "Documentação do processo de notificação de credores", "Yes", "High"),
            ("HR", "Integration plan (key personnel)", "Plano de integração (pessoal-chave)", "Yes", "Medium"),
            ("Financial", "Synergies analysis", "Análise de sinergias", "Yes", "Medium"),
        ],
    }


# ---------------------------------------------------------------------------
# Labels — bilingual
# ---------------------------------------------------------------------------

LABELS = {
    "EN": {
        "checklist_tab": "Checklist",
        "summary_tab": "Summary",
        "instructions_tab": "Instructions",
        "headers": [
            "Category", "Document Name", "Required", "Priority",
            "Received Date", "Status", "Responsible", "Comments",
        ],
        "statuses": ["Pending", "Received", "Reviewed", "Missing"],
        "summary_title": "Due Diligence — Summary",
        "target": "Target Company",
        "transaction": "Transaction Type",
        "sector": "Sector",
        "jurisdiction": "Jurisdiction",
        "date_generated": "Date Generated",
        "total_docs": "Total Documents",
        "by_category": "Documents by Category",
        "by_priority": "Documents by Priority",
        "category": "Category",
        "count": "Count",
        "priority": "Priority",
        "instructions_title": "Instructions",
        "how_to_use": "How to Use This Checklist",
        "how_to_use_items": [
            "1. Review all documents listed in the Checklist tab.",
            "2. For each document, update the Status column as you progress.",
            "3. Record the Received Date when the document is obtained.",
            "4. Assign a Responsible person for follow-up on each item.",
            "5. Use the Comments column for any observations, issues or follow-ups.",
            "6. Use the filters to focus on specific categories, priorities or statuses.",
        ],
        "status_definitions": "Status Definitions",
        "status_defs": [
            ("Pending", "Document has been requested but not yet received."),
            ("Received", "Document received but not yet reviewed by the DD team."),
            ("Reviewed", "Document reviewed; no further action needed."),
            ("Missing", "Document unavailable or target unable to provide."),
        ],
        "timeline_title": "Indicative DD Timeline",
        "timeline_items": [
            ("Week 1-2", "Send initial document request list to target / advisors."),
            ("Week 2-4", "Receive and catalogue documents in virtual data room."),
            ("Week 3-6", "Detailed review by legal, financial and tax workstreams."),
            ("Week 5-7", "Follow-up requests and Q&A with management."),
            ("Week 7-8", "Draft DD reports and identify key findings / red flags."),
            ("Week 8-10", "Final DD reports issued; feed into SPA negotiation."),
        ],
        "contacts_title": "Advisor Contacts",
        "contacts_headers": ["Role", "Firm", "Contact Person", "Email", "Phone"],
        "contacts_roles": [
            "Legal Advisor", "Financial Advisor", "Tax Advisor",
            "Environmental Advisor", "Insurance Advisor", "IT / Cyber Advisor",
        ],
    },
    "PT": {
        "checklist_tab": "Checklist",
        "summary_tab": "Resumo",
        "instructions_tab": "Instruções",
        "headers": [
            "Categoria", "Nome do Documento", "Obrigatório", "Prioridade",
            "Data de Receção", "Estado", "Responsável", "Comentários",
        ],
        "statuses": ["Pendente", "Recebido", "Revisto", "Em falta"],
        "summary_title": "Due Diligence — Resumo",
        "target": "Empresa-alvo",
        "transaction": "Tipo de Transação",
        "sector": "Setor",
        "jurisdiction": "Jurisdição",
        "date_generated": "Data de Geração",
        "total_docs": "Total de Documentos",
        "by_category": "Documentos por Categoria",
        "by_priority": "Documentos por Prioridade",
        "category": "Categoria",
        "count": "Contagem",
        "priority": "Prioridade",
        "instructions_title": "Instruções",
        "how_to_use": "Como Usar Esta Checklist",
        "how_to_use_items": [
            "1. Reveja todos os documentos listados no separador Checklist.",
            "2. Para cada documento, atualize a coluna Estado à medida que avança.",
            "3. Registe a Data de Receção quando o documento for obtido.",
            "4. Atribua um Responsável pelo acompanhamento de cada item.",
            "5. Use a coluna Comentários para observações, questões ou seguimentos.",
            "6. Utilize os filtros para focar em categorias, prioridades ou estados específicos.",
        ],
        "status_definitions": "Definições de Estado",
        "status_defs": [
            ("Pendente", "Documento solicitado mas ainda não recebido."),
            ("Recebido", "Documento recebido mas ainda não revisto pela equipa de DD."),
            ("Revisto", "Documento revisto; sem ações adicionais necessárias."),
            ("Em falta", "Documento indisponível ou o target não consegue fornecer."),
        ],
        "timeline_title": "Timeline Indicativo de DD",
        "timeline_items": [
            ("Semana 1-2", "Enviar lista inicial de pedidos de documentos ao target / assessores."),
            ("Semana 2-4", "Receção e catalogação de documentos no data room virtual."),
            ("Semana 3-6", "Revisão detalhada pelas workstreams legal, financeira e fiscal."),
            ("Semana 5-7", "Pedidos de follow-up e Q&A com a gestão."),
            ("Semana 7-8", "Elaboração de relatórios de DD e identificação de red flags."),
            ("Semana 8-10", "Relatórios finais de DD; alimentar negociação do SPA."),
        ],
        "contacts_title": "Contactos dos Assessores",
        "contacts_headers": ["Função", "Firma", "Pessoa de Contacto", "Email", "Telefone"],
        "contacts_roles": [
            "Assessor Jurídico", "Assessor Financeiro", "Assessor Fiscal",
            "Assessor Ambiental", "Assessor de Seguros", "Assessor TI / Cyber",
        ],
    },
}


# ---------------------------------------------------------------------------
# Build the full document list
# ---------------------------------------------------------------------------

def build_document_list(deal_type: str, sector: str, lang: str) -> list[tuple]:
    """Return list of (category, name, required, priority) tuples."""
    name_idx = 1 if lang == "EN" else 2
    docs: list[tuple] = []

    for doc in CORE_DOCUMENTS:
        docs.append((doc[0], doc[name_idx], doc[3], doc[4]))

    sector_docs = _sector_documents().get(sector, [])
    for doc in sector_docs:
        docs.append((doc[0], doc[name_idx], doc[3], doc[4]))

    deal_docs = _deal_documents().get(deal_type, [])
    for doc in deal_docs:
        docs.append((doc[0], doc[name_idx], doc[3], doc[4]))

    # Sort by category then priority rank
    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    docs.sort(key=lambda d: (d[0], priority_order.get(d[3], 9)))
    return docs


# ---------------------------------------------------------------------------
# Excel formatting helpers
# ---------------------------------------------------------------------------

DARK_BLUE = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12)
BODY_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
FILL_HIGH = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
FILL_MEDIUM = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_LOW = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

STATUS_FILLS = {
    "Pending": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Received": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "Reviewed": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Missing": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    # PT equivalents
    "Pendente": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Recebido": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "Revisto": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Em falta": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
}

PRIORITY_FILLS = {"High": FILL_HIGH, "Medium": FILL_MEDIUM, "Low": FILL_LOW}


def _auto_width(ws, min_width=12, max_width=55):
    """Auto-adjust column widths based on cell content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        best = min_width
        for cell in col_cells:
            if cell.value:
                best = max(best, min(len(str(cell.value)) + 3, max_width))
        ws.column_dimensions[col_letter].width = best


# ---------------------------------------------------------------------------
# Tab 1 — Checklist
# ---------------------------------------------------------------------------

def create_checklist_tab(wb: Workbook, docs: list[tuple], lang: str, labels: dict):
    ws = wb.active
    ws.title = labels["checklist_tab"]

    headers = labels["headers"]
    statuses = labels["statuses"]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = DARK_BLUE
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Write document rows
    default_status = statuses[0]  # Pending / Pendente
    for row_idx, (cat, name, req, prio) in enumerate(docs, 2):
        values = [cat, name, req, prio, "", default_status, "", ""]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 2))

        # Priority fill (column 4)
        prio_cell = ws.cell(row=row_idx, column=4)
        if prio in PRIORITY_FILLS:
            prio_cell.fill = PRIORITY_FILLS[prio]

        # Status fill (column 6)
        status_cell = ws.cell(row=row_idx, column=6)
        if default_status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[default_status]

    # Conditional formatting for Status column (F)
    last_row = len(docs) + 1
    status_range = f"F2:F{last_row}"
    for status_val, fill in STATUS_FILLS.items():
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{status_val}"'], fill=fill),
        )

    # Data validation for Status column
    dv_status = DataValidation(
        type="list",
        formula1='"' + ",".join(statuses) + '"',
        allow_blank=True,
    )
    dv_status.error = "Please select a valid status."
    dv_status.errorTitle = "Invalid Status"
    ws.add_data_validation(dv_status)
    dv_status.add(f"F2:F{last_row}")

    # Data validation for Priority column
    dv_prio = DataValidation(
        type="list", formula1='"High,Medium,Low"', allow_blank=True,
    )
    ws.add_data_validation(dv_prio)
    dv_prio.add(f"D2:D{last_row}")

    # Auto-filter & freeze
    ws.auto_filter.ref = f"A1:H{last_row}"
    ws.freeze_panes = "A2"

    _auto_width(ws)
    # Force document name column wider
    ws.column_dimensions["B"].width = 55


# ---------------------------------------------------------------------------
# Tab 2 — Summary
# ---------------------------------------------------------------------------

def create_summary_tab(
    wb: Workbook, docs: list[tuple], labels: dict,
    target: str, deal_type: str, sector: str, jurisdiction: str,
):
    ws = wb.create_sheet(title=labels["summary_tab"])

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value=labels["summary_title"])
    title_cell.font = TITLE_FONT

    # Metadata
    meta = [
        (labels["target"], target),
        (labels["transaction"], deal_type),
        (labels["sector"], sector),
        (labels["jurisdiction"], jurisdiction),
        (labels["date_generated"], datetime.now().strftime("%Y-%m-%d %H:%M")),
        (labels["total_docs"], len(docs)),
    ]
    for i, (label, value) in enumerate(meta, 3):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws.cell(row=i, column=2, value=value).font = BODY_FONT

    # Breakdown by category
    row = len(meta) + 5
    ws.cell(row=row, column=1, value=labels["by_category"]).font = SUBTITLE_FONT
    row += 1
    cat_header_lbl = ws.cell(row=row, column=1, value=labels["category"])
    cat_header_cnt = ws.cell(row=row, column=2, value=labels["count"])
    for c in (cat_header_lbl, cat_header_cnt):
        c.font = HEADER_FONT
        c.fill = DARK_BLUE
        c.border = THIN_BORDER

    cat_counts = {}
    for d in docs:
        cat_counts[d[0]] = cat_counts.get(d[0], 0) + 1

    for cat in CATEGORIES:
        count = cat_counts.get(cat, 0)
        if count > 0:
            row += 1
            ws.cell(row=row, column=1, value=cat).font = BODY_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=count).font = BODY_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER

    # Breakdown by priority
    row += 2
    ws.cell(row=row, column=1, value=labels["by_priority"]).font = SUBTITLE_FONT
    row += 1
    prio_header_lbl = ws.cell(row=row, column=1, value=labels["priority"])
    prio_header_cnt = ws.cell(row=row, column=2, value=labels["count"])
    for c in (prio_header_lbl, prio_header_cnt):
        c.font = HEADER_FONT
        c.fill = DARK_BLUE
        c.border = THIN_BORDER

    prio_counts = {}
    for d in docs:
        prio_counts[d[3]] = prio_counts.get(d[3], 0) + 1

    for prio in ("High", "Medium", "Low"):
        count = prio_counts.get(prio, 0)
        if count > 0:
            row += 1
            c1 = ws.cell(row=row, column=1, value=prio)
            c1.font = BODY_FONT
            c1.border = THIN_BORDER
            if prio in PRIORITY_FILLS:
                c1.fill = PRIORITY_FILLS[prio]
            c2 = ws.cell(row=row, column=2, value=count)
            c2.font = BODY_FONT
            c2.border = THIN_BORDER

    _auto_width(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


# ---------------------------------------------------------------------------
# Tab 3 — Instructions
# ---------------------------------------------------------------------------

def create_instructions_tab(wb: Workbook, labels: dict):
    ws = wb.create_sheet(title=labels["instructions_tab"])

    row = 1
    ws.merge_cells("A1:E1")
    ws.cell(row=row, column=1, value=labels["instructions_title"]).font = TITLE_FONT

    # How to use
    row += 2
    ws.cell(row=row, column=1, value=labels["how_to_use"]).font = SUBTITLE_FONT
    for item in labels["how_to_use_items"]:
        row += 1
        ws.cell(row=row, column=1, value=item).font = BODY_FONT

    # Status definitions
    row += 2
    ws.cell(row=row, column=1, value=labels["status_definitions"]).font = SUBTITLE_FONT
    row += 1
    for c, hdr in [(1, "Status"), (2, "Definition" if "how_to_use" in labels else "Definição")]:
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = DARK_BLUE
        cell.border = THIN_BORDER
    for status, definition in labels["status_defs"]:
        row += 1
        c1 = ws.cell(row=row, column=1, value=status)
        c1.font = BOLD_FONT
        c1.border = THIN_BORDER
        if status in STATUS_FILLS:
            c1.fill = STATUS_FILLS[status]
        c2 = ws.cell(row=row, column=2, value=definition)
        c2.font = BODY_FONT
        c2.border = THIN_BORDER

    # Timeline
    row += 2
    ws.cell(row=row, column=1, value=labels["timeline_title"]).font = SUBTITLE_FONT
    row += 1
    for c, hdr in [(1, "Phase"), (2, "Activities")]:
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = DARK_BLUE
        cell.border = THIN_BORDER
    for phase, desc in labels["timeline_items"]:
        row += 1
        ws.cell(row=row, column=1, value=phase).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).font = BODY_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

    # Contacts template
    row += 2
    ws.cell(row=row, column=1, value=labels["contacts_title"]).font = SUBTITLE_FONT
    row += 1
    for c, hdr in enumerate(labels["contacts_headers"], 1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = DARK_BLUE
        cell.border = THIN_BORDER
    for role in labels["contacts_roles"]:
        row += 1
        ws.cell(row=row, column=1, value=role).font = BODY_FONT
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER

    _auto_width(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# Terminal UI
# ---------------------------------------------------------------------------

def choose(prompt: str, options: list[str]) -> str:
    """Display a numbered menu and return the chosen option."""
    print(f"\n{prompt}")
    for i, opt in enumerate(options, 1):
        print(f"  [{i}] {opt}")
    while True:
        raw = input("  > ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(options):
            choice = options[int(raw) - 1]
            print(f"  ✔ {choice}")
            return choice
        print(f"  ✗ Please enter a number between 1 and {len(options)}.")


def ask_text(prompt: str, allow_empty=False) -> str:
    """Ask for free-text input."""
    while True:
        val = input(f"\n{prompt}: ").strip()
        if val or allow_empty:
            return val
        print("  ✗ This field cannot be empty.")


def ask_yes_no(prompt: str) -> bool:
    """Ask a yes/no question."""
    while True:
        val = input(f"\n{prompt} (y/n): ").strip().lower()
        if val in ("y", "yes", "s", "sim"):
            return True
        if val in ("n", "no", "nao", "não"):
            return False
        print("  ✗ Please answer y or n.")


def print_preview(docs: list[tuple], labels: dict):
    """Print a preview of the checklist to the terminal."""
    headers = labels["headers"]
    print("\n" + "=" * 90)
    print(f"  {'PREVIEW':^86}")
    print("=" * 90)
    print(f"  {headers[0]:<14} {headers[1]:<46} {headers[2]:<10} {headers[3]}")
    print("-" * 90)
    for cat, name, req, prio in docs:
        name_display = (name[:43] + "...") if len(name) > 46 else name
        print(f"  {cat:<14} {name_display:<46} {req:<10} {prio}")
    print("-" * 90)
    print(f"  Total: {len(docs)} documents")
    print("=" * 90)


def ask_custom_documents(lang: str) -> list[tuple]:
    """Allow user to add custom documents interactively."""
    custom: list[tuple] = []
    prompt_cat = "Category" if lang == "EN" else "Categoria"
    prompt_name = "Document name" if lang == "EN" else "Nome do documento"
    prompt_req = "Required? (y/n)" if lang == "EN" else "Obrigatório? (s/n)"
    prompt_more = "Add another document?" if lang == "EN" else "Adicionar outro documento?"

    while True:
        print()
        cat = choose(f"  {prompt_cat}:", CATEGORIES)
        name = ask_text(f"  {prompt_name}")
        req = "Yes" if ask_yes_no(f"  {prompt_req}") else "No"
        prio = choose("  Priority:", ["High", "Medium", "Low"])
        custom.append((cat, name, req, prio))
        print(f"  ✔ Added: {name}")
        if not ask_yes_no(f"  {prompt_more}"):
            break
    return custom


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate_excel(
    target: str,
    deal_type: str,
    sector: str,
    jurisdiction: str,
    lang: str,
    docs: list[tuple],
) -> str:
    """Generate the Excel file and return the output path."""
    wb = Workbook()
    labels = LABELS[lang]

    create_checklist_tab(wb, docs, lang, labels)
    create_summary_tab(wb, docs, labels, target, deal_type, sector, jurisdiction)
    create_instructions_tab(wb, labels)

    safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in target).strip()
    safe_name = safe_name.replace(" ", "_")
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"{safe_name}_DD_Checklist_{date_str}.xlsx"

    wb.save(filename)
    return filename


def run_interactive():
    """Run the interactive terminal flow."""
    print("\n" + "=" * 60)
    print("   DUE DILIGENCE — DOCUMENT CHECKLIST GENERATOR")
    print("=" * 60)

    lang = choose("Language / Idioma:", ["EN — English", "PT — Português"])
    lang = "EN" if lang.startswith("EN") else "PT"

    deal_type = choose(
        "Transaction type:" if lang == "EN" else "Tipo de transação:",
        TRANSACTION_TYPES,
    )
    sector = choose("Sector:" if lang == "EN" else "Setor:", SECTORS)
    jurisdiction = choose(
        "Jurisdiction:" if lang == "EN" else "Jurisdição:",
        JURISDICTIONS,
    )
    target = ask_text("Target company name" if lang == "EN" else "Nome da empresa-alvo")

    # Build document list
    docs = build_document_list(deal_type, sector, lang)

    # Preview
    labels = LABELS[lang]
    print_preview(docs, labels)

    # Custom documents
    if ask_yes_no(
        "Add custom documents?" if lang == "EN" else "Adicionar documentos personalizados?"
    ):
        custom = ask_custom_documents(lang)
        docs.extend(custom)
        print(f"\n  → {len(custom)} custom document(s) added. New total: {len(docs)}")

    # Generate
    if ask_yes_no("Generate Excel file?" if lang == "EN" else "Gerar ficheiro Excel?"):
        filepath = generate_excel(target, deal_type, sector, jurisdiction, lang, docs)
        abs_path = str(Path(filepath).resolve())
        print("\n" + "=" * 60)
        print(f"  ✔ File generated: {abs_path}")
        print("=" * 60)
        return abs_path
    else:
        print("\n  Cancelled.")
        return None


def run_automated(
    target: str,
    deal_type: str,
    sector: str,
    jurisdiction: str,
    lang: str = "EN",
    custom_docs: list[tuple] | None = None,
) -> str:
    """Non-interactive entry point for testing / integration."""
    if deal_type not in TRANSACTION_TYPES:
        raise ValueError(f"Invalid deal type: {deal_type}. Must be one of {TRANSACTION_TYPES}")
    if sector not in SECTORS:
        raise ValueError(f"Invalid sector: {sector}. Must be one of {SECTORS}")
    if jurisdiction not in JURISDICTIONS:
        raise ValueError(f"Invalid jurisdiction: {jurisdiction}. Must be one of {JURISDICTIONS}")
    if lang not in ("EN", "PT"):
        raise ValueError(f"Invalid language: {lang}. Must be EN or PT")

    docs = build_document_list(deal_type, sector, lang)
    if custom_docs:
        docs.extend(custom_docs)

    return generate_excel(target, deal_type, sector, jurisdiction, lang, docs)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--test":
        # Quick automated test
        print("Running automated test...")
        path = run_automated(
            target="TechVida Lda",
            deal_type="Share Deal",
            sector="Technology",
            jurisdiction="Portugal",
            lang="EN",
        )
        print(f"Test file generated: {path}")

        path_pt = run_automated(
            target="Farma Saúde SA",
            deal_type="Merger",
            sector="Healthcare",
            jurisdiction="Portugal",
            lang="PT",
        )
        print(f"Test file (PT) generated: {path_pt}")
    else:
        run_interactive()
